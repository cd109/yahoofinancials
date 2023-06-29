from yahoofinancials import YahooFinancials
import pandas as pd
import openpyxl


def GET_annual_FILE(dataType , stockCode , stockName , currency , financial_stmts):
    sheet_name = stockCode + '-' + currency + '-' + stockName
    yahoo_financials = YahooFinancials(stockCode)
    if dataType == 'IS': statement_data_qt = yahoo_financials.get_financial_stmts('annual', 'income' )
    if dataType == 'CF': statement_data_qt = yahoo_financials.get_financial_stmts('annual', 'cash' )
    if dataType == 'BS': statement_data_qt = yahoo_financials.get_financial_stmts('annual', 'balance' )
    dict_list = statement_data_qt[financial_stmts][stockCode]
    df = pd.concat([pd.DataFrame(i) for i in dict_list], axis=1)
    df = df.rename_axis(financial_stmts).reset_index()
    df.insert(0, 'ticker', stockCode)
    df.drop(df.columns[[0]], axis=1, inplace=True)
    file_name = 'D:\\ALL_2\\yahoofinancials\\FIN_0617\\yahooExport_annual_'+ dataType + '_' + stockCode + '.xlsx'
    df.to_excel(file_name , sheet_name = sheet_name)
    
    wk = openpyxl.load_workbook(file_name) 
    wk_sheet = wk[sheet_name]
    wk_sheet.delete_cols(1)
    wk_sheet.cell(1,1,value='Breakdown')
    wk.save(file_name)



def GET_quarterly_FILE(dataType , stockCode , stockName , currency , financial_stmts):
    sheet_name = stockCode + '-' + currency + '-' + stockName
    yahoo_financials = YahooFinancials(stockCode)
    if dataType == 'IS': statement_data_qt = yahoo_financials.get_financial_stmts('quarterly', 'income' )
    if dataType == 'CF': statement_data_qt = yahoo_financials.get_financial_stmts('quarterly', 'cash' )
    if dataType == 'BS': statement_data_qt = yahoo_financials.get_financial_stmts('quarterly', 'balance' )
    dict_list = statement_data_qt[financial_stmts][stockCode]

    if len(dict_list) > 0 :
       df = pd.concat([pd.DataFrame(i) for i in dict_list], axis=1)
       df = df.rename_axis(financial_stmts).reset_index()
       df.insert(0, 'ticker', stockCode)
       df.drop(df.columns[[0]], axis=1, inplace=True)
       file_name = 'D:\\ALL_2\\yahoofinancials\\FIN_0617\\yahooExport_Quarterly_'+ dataType + '_' + stockCode + '.xlsx'
       df.to_excel(file_name , sheet_name = sheet_name)

       wk = openpyxl.load_workbook(file_name) 
       wk_sheet = wk[sheet_name]
       wk_sheet.delete_cols(1)
       wk_sheet.cell(1,1,value='Breakdown')
       wk.save(file_name)
    else:   
       print(stockCode + " is pass !!")